from fileinput import filename

import pandas as pd
from openpyxl import load_workbook
import requests
import tempfile
from pathlib import Path
import zipfile
import time



def read_policy_file(uploaded_file):

    df = pd.read_excel(
        uploaded_file,
        sheet_name="Policy Details"
    )

    return df


def get_policy_summary(df):

    total_policies = len(df)

    charge_col = None

    for col in df.columns:

        col_lower = str(col).lower()

        if (
            "charge" in col_lower
            or "premium" in col_lower
            or "amount" in col_lower
        ):
            charge_col = col
            break

    if not charge_col:
        return total_policies, 0

    charges = (
        df[charge_col]
        .fillna("")
        .astype(str)
        .str.replace(r"[^\d.\-]", "", regex=True)
        .str.strip()
    )

    total_charges = float(pd.to_numeric(
        charges,
        errors="coerce"
    ).fillna(0).sum())

    return total_policies, total_charges

def read_policy_links(uploaded_file):

    wb = load_workbook(
        uploaded_file,
        data_only=True
    )

    ws = wb["Policy Details"]

    headers = {}

    for col in range(1, ws.max_column + 1):
        headers[
            ws.cell(1, col).value
        ] = col

    records = []

    hub_column = None

    for header in headers:
        if str(header).strip().lower() in [
            "hub",
            "hub name",
            "hub_name"
        ]:
            hub_column = header
            break

    for row in range(2, ws.max_row + 1):

        policy_no = ws.cell(
            row,
            headers["Policy Number"]
        ).value

        traveller = ws.cell(
            row,
            headers["Traveller Name"]
        ).value

        hub_name = ""

        if hub_column:
            hub_name = ws.cell(
                row,
                headers[hub_column]
            ).value

        link_cell = ws.cell(
            row,
            headers["Policy Path"]
        )

        url = ""

        if link_cell.hyperlink:
            url = link_cell.hyperlink.target

        records.append({
            "policy_number": policy_no,
            "traveller_name": traveller,
            "url": url,
            "hub_name": hub_name
        })

    return pd.DataFrame(records)

def clean_filename(text):

    invalid_chars = r'\/:*?"<>|'

    text = str(text)

    for ch in invalid_chars:
        text = text.replace(ch, "_")

    return text.strip()

def download_policy_pdfs(
    link_df,
    progress_callback=None,
    hub_wise=False
    ):

    base_dir = Path(
        tempfile.mkdtemp()
    )

    first_traveller = clean_filename(
        str(link_df.iloc[0]["traveller_name"])
    )

    first_policy = clean_filename(
        str(link_df.iloc[0]["policy_number"])
    )

    folder_name = (
        f"{first_traveller} [{first_policy}]"
    )

    temp_dir = base_dir / folder_name

    temp_dir.mkdir(
        parents=True,
        exist_ok=True
    )

    downloaded = 0
    failed = 0

    start_time = time.time()
    total = len(link_df)

    hub_counts = {}

    if hub_wise:
        for hub in link_df["hub_name"].fillna(""):

            hub = str(hub).strip()

            if not hub:
                hub = "Unknown Hub"

            hub_counts[hub] = (
                hub_counts.get(hub, 0) + 1
            )

    for idx, (_, row) in enumerate(
        link_df.iterrows(),
        start=1
    ):

        policy_no = str(
            row["policy_number"]
        ).strip()

        traveller = str(
            row["traveller_name"]
        ).strip()

        url = str(
            row["url"]
        ).strip()

        if not url:

            failed += 1

            elapsed = time.time() - start_time

            if progress_callback:

                progress_callback(
                    idx,
                    total,
                    elapsed,
                    0
                )

            continue

        filename = (
            clean_filename(
                f"{traveller} [{policy_no}]"
            )
            + ".pdf"
        )

        if hub_wise:

            hub_name = str(
                row.get("hub_name", "")
            ).strip()

            if not hub_name:
                hub_name = "Unknown Hub"

            folder_display_name = (
                f"{hub_name}_{hub_counts[hub_name]}_Policies"
            )

            hub_folder = (
                temp_dir /
                clean_filename(folder_display_name)
            )

            hub_folder.mkdir(
                parents=True,
                exist_ok=True
            )

            filepath = hub_folder / filename

        else:

            filepath = temp_dir / filename

        try:

            response = requests.get(
                url,
                timeout=60
            )

            if response.status_code == 200:

                with open(
                    filepath,
                    "wb"
                ) as f:

                    f.write(response.content)

                downloaded += 1

            else:

                failed += 1

        except Exception:

            failed += 1

        elapsed = time.time() - start_time

        if progress_callback:

            progress_callback(
                idx,
                total,
                elapsed,
                0
            )

    if len(link_df) == 1:
        zip_name = f"{folder_name}.zip"
    else:
        zip_name = (
            f"{folder_name} x {len(link_df)}.zip"
        )

    zip_path = base_dir / zip_name

    with zipfile.ZipFile(
        zip_path,
        "w",
        zipfile.ZIP_DEFLATED
    ) as zipf:

        for pdf_file in temp_dir.rglob("*.pdf"):

            zipf.write(
                pdf_file,
                arcname=str(
                    pdf_file.relative_to(base_dir)
                )
            )

    total_time = time.time() - start_time

    return zip_path, downloaded, failed, total_time
